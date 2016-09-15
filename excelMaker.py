import pathCreator
import os
import pandas as pd
import openpyxl as op


def makeDataFrame(appendedWorkDir, order):
	counterDict = {}
	with open(os.path.join(appendedWorkDir, 'D{}ResultsDict.txt'.format(order)), 'r') as f:
		counterDict = eval(f.read())
	return pd.DataFrame(counterDict)


def sortedDataFrame(appendedWorkDir, order):
	orderList = []

	with open(os.path.join(appendedWorkDir, 'D{}Group.txt'.format(order)), 'r') as f:
		orderList = eval(f.read())

	myFrame = makeDataFrame(appendedWorkDir, order).reindex_axis(orderList, axis=1)
	myFrame = myFrame.reindex_axis(orderList, axis=0)

	return myFrame



def makeExcelSheet(myWorkDir, appendedWorkDir, order):
	groupFrame = sortedDataFrame(appendedWorkDir, order)

	if not (os.path.exists('DihedralGroups.xlsx')):
		wb = op.Workbook()
		wb.save('DihedralGroups.xlsx')

	
	book = op.load_workbook('DihedralGroups.xlsx')
	writer = pd.ExcelWriter('DihedralGroups.xlsx', engine='openpyxl') 
	writer.book = book
	writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

	groupFrame.to_excel(writer, "D{}".format(order))
	writer.save()


def colorCells(myWorkDir, order):
	completeFill = op.styles.PatternFill(start_color='FF77CAE6', end_color='FF77CAE6', fill_type='solid')
	anomalyFill = op.styles.PatternFill(start_color='FFE69377', end_color='FFE69377', fill_type='solid')

	wb = op.load_workbook('DihedralGroups.xlsx')

	ws = wb.get_sheet_by_name("D{}".format(order))

	for row in ws.rows:
		for cell in row:
			if isinstance(cell.value, int):
				if cell.value == order:
					cell.fill = completeFill
				if pathCreator.isAnomalous(cell.value, order):
					cell.fill = anomalyFill

	wb.save('DihedralGroups.xlsx')


def mainFunction(order):
	myWorkDir = os.path.dirname(__file__)
	appendedWorkDir = pathCreator.writeResults(order)
	makeExcelSheet(myWorkDir, appendedWorkDir, order)
	colorCells(myWorkDir, order)


if __name__ == '__main__':
	order = int(input('2n = ? '))
	mainFunction(order)

